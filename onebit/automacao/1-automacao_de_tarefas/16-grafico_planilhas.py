# Criando gráfico a partir de duas planilhas

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference, series

dict_anos = {}

# 1 -> Importando despesas
arquivo1 = load_workbook(filename='files/despesas.xlsx')
ws1 = arquivo1['despesas']
max_linhas = ws1.max_row

for i in range(2, max_linhas+1):
    dict_anos[ws1['A%s' %i].value] = {'despesas': ws1['B%s' %i].value, 'receita':0}

# 2 -> Importando receitas
arquivo1 = load_workbook(filename='files/receita.xlsx')
ws2 = arquivo1['receita']
max_linhas = ws2.max_row

for i in range(2,max_linhas+1):
    dict_anos[ws2['A%s'%i].value]['receita'] = ws2['B%s'%i].value

# 3 -> Criando a planilha
wb = Workbook()
ws = wb.active
ws.title = 'Demonstrativo'

ws['A1'] = 'Ano'
ws['B1'] = 'Despesas'
ws['C1'] = 'Receita'

i = 2 # -> Para direcionar qual linha irá pegar, exemplo a2, a3, pois assim não irá pegar o título da planilha.  

for key,values in dict_anos.items():
    ws['A%s'%i] = key
    ws['B%s'%i] = values['despesas']
    ws['C%s'%i] = values['receita']
    i += 1

# 4 -> Criando gráfico
chart1 = BarChart()
chart1.type = 'col'
chart1.style = 10
chart1.title = "Receita x Despesas por ano"
chart1.y_axis.title = 'R$'
chart1.x_axis.title = 'Ano'

# Passando os dados para o gráfico
data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=i)

anos = Reference(ws, min_col=1, min_row=2, max_row=i)

chart1.add_data(data,titles_from_data=True)
chart1.set_categories(anos)
chart1.shape = 4

ws.add_chart(chart1,'A%s' %(i+2))
wb.save(filename='files/demonstrativo.xlsx')